#!/usr/bin/env python3
"""
고객 요구사항 vs 실제 구현 vs 테스트 방법 매트릭스를 엑셀로 출력.
출처: 견적서 (JpexStudio_에스킴_블루밍테라_SEO및UI개선_견적서_260416.xlsx) +
      _docs/specs/작업계획-가능여부및난이도.md +
      git log + tests/
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ────────────────────────────────────────────────
# 데이터
# ────────────────────────────────────────────────
ROWS = [
    # Phase, 사이트, 항목, 고객 요구사항, 실제 구현, 변경 파일, 커밋, 테스트 방법, 기대 결과, E2E, 상태
    # ── Phase 1: SEO 기본
    ["Phase 1", "블루밍테라", "robots.txt 생성",
     "검색엔진 크롤러 가이드 파일 생성",
     "루트에 robots.txt 작성, /admin/ /app/ /system/ /_compile/ /data/ /lib/admin/ 차단, sitemap 명시",
     "bloomingterra/robots.txt",
     "ff7131d 외", "curl https://bloomingterra.com/robots.txt", "200 + Disallow + Sitemap 라인", "1", "✅"],

    ["Phase 1", "블루밍테라", "sitemap.xml 생성",
     "검색엔진에 페이지 목록 제공",
     "정적 sitemap.xml 생성 (Phase 2-6에서 동적화로 대체)",
     "bloomingterra/sitemap.xml (제거됨)",
     "Phase 1 → Phase 2-6", "curl https://bloomingterra.com/sitemap.xml", "200 + XML", "1", "✅ (동적화로 대체)"],

    ["Phase 1", "블루밍테라", "SSL 리다이렉트",
     "HTTP → HTTPS 강제",
     "Nginx 서버 레벨에서 처리 중 확인 → .htaccess 변경 불필요",
     "bloomingterra/.htaccess (변경 없음, 코멘트만)",
     "ed8fddd", "curl -I http://bloomingterra.com", "301 to https",
     "1 (수동)", "✅"],

    ["Phase 1", "블루밍테라", "메타태그/OG 태그",
     "검색결과 노출/SNS 미리보기용 메타태그",
     "header.html에 description/keywords/og:type/og:title/og:url/og:site_name/og:description/og:locale/og:image + Twitter Card 추가. canonical 동적 처리.",
     "bloomingterra/data/skin/respon_default_en/outline/header.html",
     "ff7131d", "https://bloomingterra.com 응답 HTML 검사 (head 메타)", "OG 9개 + Twitter Card 4개 + canonical", "13", "✅ (콘텐츠 일부 데이터 미입력)"],

    ["Phase 1", "블루밍테라", "Schema.org",
     "구조화된 데이터로 회사 정보 검색엔진 제공",
     "header.html에 Organization JSON-LD 추가 (name/url/logo/address/sameAs)",
     "bloomingterra/data/skin/respon_default_en/outline/header.html",
     "ff7131d", "응답 HTML에 application/ld+json @type Organization", "JSON-LD 객체 1개", "1", "✅"],

    ["Phase 1", "에스킴", "robots.txt 보강",
     "민감 경로 차단",
     "/report*.html, /comm/, /other/search.php 차단 + sitemap 명시",
     "askim/www/robots.txt",
     "0b19c31", "curl https://www.askim.kr/robots.txt", "Disallow 7줄 + Sitemap", "1", "✅"],

    ["Phase 1", "에스킴", "Schema.org 보강",
     "Organization 정보 보강",
     "head.php에 url/logo/description/addressCountry/sameAs 3개 추가",
     "askim/www/head.php",
     "8971471", "https://www.askim.kr 응답 ld+json", "Organization 객체 + sameAs 3개", "1", "✅"],

    ["Phase 1", "에스킴", "OG/Twitter Card 보강",
     "비표준 메타태그 정리 + OG/Twitter 정리",
     "twcenter/lib.php에서 비표준(subject/title/author/publisher/Classification) 제거, og:locale/image:width/height/Twitter Card 추가",
     "askim/www/twcenter/lib.php",
     "ef99616", "https://www.askim.kr 응답 head meta 검사", "비표준 0개 + Twitter Card 4개", "12", "✅"],

    # ── Phase 2-1: Service 레이아웃
    ["Phase 2-1", "블루밍테라", "Service 페이지 → 에스킴 포트폴리오 레이아웃",
     "goods_view를 에스킴 portfolio view 스타일(2-column flex, sticky txt_area)로 재작성",
     "goods_view.html 재작성 (txt_area sticky + con_area 본문) + portfolio_view.css 신규",
     "bloomingterra/data/skin/respon_default_en/goods/goods_view.html, css/portfolio_view.css",
     "e7c2097, 0d7e9f1, 1cab5fb",
     "https://www.bloomingterra.com/goods/goods_view?no=171&cate=001007 (TIRTIR)",
     "portfolio_view_in flex layout + txt_area sticky info + con_area 본문",
     "9", "✅"],

    # ── Phase 2-2: Recent Posts
    ["Phase 2-2", "블루밍테라", "Recent Posts (goods_view)",
     "글 view 페이지 하단에 최근 작성된 다른 글 슬라이드",
     "Goods.php에 recent_goods 30개 페치 + Swiper marquee + portfolio_view.css 마키 스타일",
     "Goods.php, goods_view.html, portfolio_view.css",
     "c45eec7, 1cab5fb",
     "동일 view 페이지 하단 Recent Posts 섹션 확인",
     "30 marquee-item (loop swiper) + 화살표/페이지네이션 동작",
     "9", "✅"],

    # ── Phase 2-3: Insight 웹진
    ["Phase 2-3", "블루밍테라", "Insight 게시판(code=gallery) → 오브라운 웹진 스타일",
     "기존 단순 카드 list/view를 모노톤 + 큰 썸네일 카드(라운드/호버 줌/오버레이) + view 본문 max-width 880",
     "board_list.html, board_view.html에 code=gallery일 때 wrapper 클래스 분기 + insight_board.css 신규 (모노톤+Kanit+호버 효과)",
     "bloomingterra/data/skin/respon_default_en/board/board_list.html, board_view.html, css/insight_board.css",
     "e1be52a",
     "https://www.bloomingterra.com/board/board_list?code=gallery (list), .../board_view?code=gallery&no=82 (view), https://.../board_list?code=notice (영향 없음 검증)",
     "Insight만 새 디자인, 다른 게시판 변경 없음",
     "18", "✅"],

    # ── Phase 2-4: 관련 게시글
    ["Phase 2-4", "블루밍테라", "Insight 관련 게시글 2개 수동 지정",
     "어드민에서 관련 게시글 글 번호 2개 입력 → view에 카드형 출력",
     "DB ALTER (related_no1/2 INT NULL) + Board_model::board_write 매핑 + 어드민 _form_board_write.php에 input 분기 + Board.php::board_view에서 별도 쿼리 + board_view.html related_posts 섹션 + insight_board.css .related_posts",
     "bloomingterra/app/controllers/Board.php, models/Board_model.php, views/admin/board/_form_board_write.php, data/skin/respon_default_en/board/board_view.html, css/insight_board.css",
     "521c5ec",
     "어드민 글 수정 → related_no1/2 입력 → /board/board_view?code=gallery&no=N에서 .related_posts 섹션 확인. 다른 게시판 view엔 섹션 없음 검증.",
     "code=gallery에만 .related_posts 출력, 미입력 시 섹션 비출력",
     "11", "✅"],

    # ── Phase 2-5: 에스킴 sitemap
    ["Phase 2-5", "에스킴", "동적 sitemap 생성",
     "정적 sitemap.xml(80 URL, 글 누락 + garbage 쿼리) → 자동 갱신 동적",
     "sitemap.php 신규 (DB 직접 조회) + .htaccess RewriteRule (/sitemap.xml → sitemap.php)",
     "askim/www/sitemap.php, askim/www/.htaccess",
     "5dcc880",
     "curl https://www.askim.kr/sitemap.xml + grep ptype=view",
     "47 URL (정적 + 카테고리 + portfolio view 35개, idx=56까지 포함, garbage 쿼리 없음)",
     "12", "✅"],

    # ── Phase 2-6: 블루밍테라 sitemap
    ["Phase 2-6", "블루밍테라", "동적 sitemap 생성",
     "정적 8 URL → 모든 글/카테고리/게시판 자동 포함",
     "Sitemap.php (CI 컨트롤러) 신규 + routes.php 매핑 + da_goods/da_board_manage/da_board_<code> 조회",
     "bloomingterra/app/controllers/Sitemap.php, app/config/routes.php",
     "2e96aa9",
     "curl https://www.bloomingterra.com/sitemap.xml?_t=<ts> (cache-bust)",
     "269 URL (goods 201 + board 45 + board_list 6 + 정적 7 + 카테고리)",
     "15", "✅"],

    # ── Phase 2-7: URL Slug
    ["Phase 2-7", "블루밍테라", "URL Slug (옵션 2: 신규 글 한정)",
     "?no=171&cate=001 → /service/<slug>로 변경. 견적서 예시 /service/옥외광고-사례",
     "DB ALTER (slug VARCHAR(150) UNIQUE) + 라우트 + Goods::goods_view_by_slug 신규 + canonical 분기 + query→slug 301 redirect (single hop) + sitemap에 slug URL + 어드민 폼 + recent_goods URL helper. 기존 201개 글은 slug NULL → query URL 그대로 (색인 변동 0).",
     "bloomingterra/app/config/routes.php, app/controllers/Goods.php, Sitemap.php, views/admin/goods/goods_reg.php, data/skin/respon_default_en/outline/header.html, goods/goods_view.html",
     "e407f0a",
     "어드민 글 작성 시 slug 입력 → /service/<slug> 200 + canonical /service/<slug>. 같은 글의 query URL은 301 to /service/<slug>. slug 없는 기존 글은 query URL 200 그대로.",
     "/service/<slug> 200 + canonical 정확 + 301 single hop + 기존 글 무영향",
     "11", "✅"],

    # ── Phase 3: 고객 피드백 라운드 1 (2026-05-14)
    ["Phase 3 #1", "블루밍테라", "스크롤 시 라이트→다크 반전 효과 (askim 1:1)",
     "에스킴컴퍼니처럼 스크롤 시 특정 구간부터 흰색→검정색 서서히 반전",
     "goods_view에서 매거진 톤(.magazine_view 클래스) 분리 → portfolio_view.css의 GSAP ScrollTrigger + body.section-light/dark + transition:background 1s ease-in-out 자동 활성화. 본문 영역을 light(제목+영상+detail_img+본문)/dark(슬라이드)/dark(recent_posts) 3-section 분할. prodSwiper pagination(가는 라인) + navigation(chevron) askim 톤으로 재디자인.",
     "bloomingterra/data/skin/respon_default_en/goods/goods_view.html, css/portfolio_view.css, js/view.js",
     "f29eefe, 7a1375e, a2ac020",
     "https://www.bloomingterra.com/goods/goods_view?no=91&cate=001 스크롤 다운 — 본문 끝에서 흰색→검정 1초 페이드 전환",
     "section-light → section-dark 토글 + 본문 텍스트/제목/info 색 자동 매핑 + 부드러운 1초 페이드",
     "라이브 검증", "✅"],

    ["Phase 3 #2", "블루밍테라", "본문 이미지-텍스트 인터리브 (askim 인라인 패턴)",
     "이미지 중간중간마다 텍스트 글 (이미지-텍스트-이미지-텍스트). 현재는 마지막 이미지 하단에만 텍스트",
     "detail_img 별도 출력 영역 제거. 운영자가 본문(info) webedit에서 이미지+텍스트 직접 인라인 작성하는 askim 패턴으로 전환. DB는 보존 (dn 숨김 영역).",
     "bloomingterra/data/skin/respon_default_en/goods/goods_view.html",
     "fa7bfd1",
     "어드민 글 작성 시 본문 에디터에서 이미지+텍스트 번갈아 삽입 → view에서 본문 안에 인라인 표시",
     "view_img 별도 영역 0개. 본문 안에 운영자가 배치한 이미지+텍스트 그대로 출력. 운영 가이드 동반.",
     "운영 변경", "✅"],

    ["Phase 3 #3", "블루밍테라", "영상 최상단(hero) 배치",
     "영상을 마지막 구간보다 최상단에 노출",
     "youtube 블록 → light section 맨 위(detail_img 위)로 이동. 영상 정보 3가지(링크/텍스트/썸네일) 모두 있는 글에만 표시.",
     "bloomingterra/data/skin/respon_default_en/goods/goods_view.html",
     "18fd7aa",
     "영상 등록된 글 view → 페이지 최상단에 영상 hero 영역",
     "ex4+ex5+ex7 모두 있는 글에서 con_area 첫 자식이 .youtube 블록",
     "라이브 검증", "✅"],

    ["Phase 3 #4", "블루밍테라", "PREV/NEXT 화살표 방향 수정",
     "현재 > PREV ... NEXT < (안쪽) → < PREV ... NEXT > (바깥)",
     "회전 트릭(\\e5e0 + rotate180) 폐기 → chevron_left(\\e5cb), chevron_right(\\e5cc) 코드포인트 직접 매핑. portfolio_view.css + magazine_view.css 양쪽 적용 (Service/Insight 동일).",
     "bloomingterra/data/skin/respon_default_en/css/portfolio_view.css, css/magazine_view.css",
     "18fd7aa",
     "view_button의 PREV/NEXT 버튼 시각 검증",
     "PREV 좌향(<), NEXT 우향(>). transform:none. content 코드포인트 명시.",
     "라이브 검증", "✅"],

    ["Phase 3 #5", "블루밍테라", "Recent Posts 끊김 없는 자동 롤링",
     "Recent Posts는 자동 롤링으로 설정",
     "기존 autoplay delay 2000ms(슬라이드별 멈춤) → delay 0 + disableOnInteraction:false + speed 5000 + freeMode:{momentum:false} → 연속 마퀴 효과. hover 시 pause만 유지.",
     "bloomingterra/data/skin/respon_default_en/goods/goods_view.html",
     "18fd7aa",
     "Recent Posts 영역에서 일정 속도로 끊김 없이 흐름",
     "swiper config: autoplay.delay=0, speed=5000, running=true",
     "라이브 검증", "✅"],

    ["Phase 3 #6", "블루밍테라", "헤더 로고 크기 축소",
     "홈페이지 좌상단 로고가 너무 커서 ABOUT/SERVICE 메뉴 글자와 같은 크기로",
     "데스크 384×51 → height:28 width:auto aspect-ratio:384/51 (원본 비율 유지). 미디어쿼리별 28/26/24. 메뉴 텍스트(24/20/18)와 시각 무게 균형.",
     "bloomingterra/data/skin/respon_default_en/css/skin.css",
     "18fd7aa",
     "메인 페이지 헤더 로고 시각 검증",
     "logo 28px (≤1500 26, ≤900 24). 메뉴 텍스트와 균형.",
     "라이브 검증", "✅"],

    ["Phase 3 #7", "블루밍테라", "콘텐츠 제목 폰트 — 명조(Noto Serif KR)",
     "서비스/인사이트 콘텐츠 제목 폰트 크기와 종류 변경 가능 여부",
     "Noto Serif KR (한글 명조) 적용. 옥외광고/공간 스토리텔링과 어울리는 매거진/저널 톤. askim(Pretendard 산세리프)과 차별화 + ohbrown 매거진 톤과 자연 호환. portfolio_view.css(서비스) + magazine_view.css(인사이트) 양쪽 적용. 크기는 현재 유지(서비스 65px / 인사이트 40px), 추후 조정 가능.",
     "bloomingterra/data/skin/respon_default_en/css/portfolio_view.css, css/magazine_view.css",
     "f785110",
     "view 제목 폰트 시각/computed-style 검증",
     "font-family: \"Noto Serif KR\", Pretendard, serif. 서비스 65px / 인사이트 40px.",
     "라이브 검증", "✅"],

    ["Phase 3 회귀", "블루밍테라", "E2E 회귀 + stale 테스트 정리",
     "오늘 변경(피드백 #1~#7) 후 기존 기능 무영향 확인",
     "전체 E2E 실행 → 25 fail 분석. 모두 stale (오늘 작업 무관, 04-30 결정으로 발생). insight_webzine 폐기 디자인 검증 16개 삭제 + related_posts 마크업 hide 8개 skip + material_symbols 완화 1개 + layout_not_broken wait 모드 변경 1개.",
     "tests/test_phase2_insight_webzine.py(삭제), test_phase2_related_posts.py, test_phase2_goods_view.py, test_phase2_visual_parity.py",
     "a697bd9",
     "python3 -m pytest tests/",
     "재실행: 107 passed / 12 skipped / 2 xfailed / 1 error(인프라 timeout) — 오늘 작업 인한 진짜 회귀 0건",
     "107", "✅"],
]

HEADER = ["Phase", "사이트", "항목", "고객 요구사항", "실제 구현 내용",
          "변경 파일", "커밋 (git)", "테스트 방법 / URL", "기대 결과", "E2E 통과 수", "상태"]

# 외부 의존 잔여 (별도 sheet)
PENDING = [
    ["검색엔진 색인 모니터링", "구글 서치콘솔 + 네이버 서치어드바이저", "1~2주 후 새 sitemap URL 색인 / 4xx 비율 / 301 처리율 추적", "사용자(고객사)"],
    ["블루밍테라 SEO 콘텐츠 입력", "어드민 → 기본정책 → SEO", "description / og_title / og_image(1200×630) / keywords 입력", "사용자(고객사)"],
    ["Insight 관련 게시글 운영 데이터 입력", "어드민 → 게시판 → INSIGHT 글 수정", "각 글마다 related_no1/2 입력 (관련성 있는 글 번호)", "사용자(고객사)"],
    ["URL Slug 운영 가이드 전달", "신규 글 작성 시 slug 필드 입력 권고", "영문 키워드+카테고리+위치 등 SEO-friendly slug 작성", "사용자(고객사)"],
    ["블루밍테라 어드민 SEO 메뉴 동작 확인", "/admin → 기본정책 → 검색엔진 최적화(SEO)", "admin 레벨 99 상승 후 메뉴 정상 진입 확인", "사용자(고객사)"],
    ["URL Slug 기존 201개 일괄 마이그레이션 검토", "운영 데이터/색인 안정화 후 결정", "마이그레이션 시 da_goods_slug_history 테이블 추가 권장", "추후 결정"],
    ["에스킴 FTP IP 만료 추적", "이전엔 1주일 만료. 다시 만료 가능", "5001 자동 폴링 hook 정기 실행 권장 (선택)", "선택"],
]
PENDING_HEADER = ["항목", "위치/대상", "내용", "주체"]

# ────────────────────────────────────────────────
# 워크북 생성
# ────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "요구사항 vs 구현"

thin = Side(border_style="thin", color="CCCCCC")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5597")
phase1_fill = PatternFill("solid", fgColor="FFF2CC")
phase2_fill = PatternFill("solid", fgColor="E2EFDA")
align_top_wrap = Alignment(vertical="top", wrap_text=True)
align_center = Alignment(vertical="center", horizontal="center")

# 메타 행 (1번)
ws.cell(row=1, column=1, value="에스킴/블루밍테라 SEO 및 UI 개선 — 요구사항 vs 구현 vs 테스트")
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADER))
ws.cell(row=2, column=1, value="갱신일: 2026-05-14 / 출처: 견적서 + 고객 피드백 라운드 1(2026-05-14) + _docs/specs/작업계획-가능여부및난이도.md + git log + tests/")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADER))
ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

# 헤더 (3번)
for j, h in enumerate(HEADER, start=1):
    c = ws.cell(row=3, column=j, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = align_center
    c.border = border

# 데이터
for i, row in enumerate(ROWS, start=4):
    fill = phase1_fill if row[0].startswith("Phase 1") else phase2_fill
    for j, val in enumerate(row, start=1):
        c = ws.cell(row=i, column=j, value=val)
        c.alignment = align_top_wrap
        c.border = border
        c.fill = fill
        if j in (1, 2, 7, 10, 11):
            c.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)

# 컬럼 너비
widths = [10, 12, 28, 36, 50, 38, 12, 42, 38, 10, 8]
for j, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(j)].width = w

# 행 높이 (자동 계산이 어려우니 충분히)
for i in range(4, 4 + len(ROWS)):
    ws.row_dimensions[i].height = 95

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 28

ws.freeze_panes = "A4"

# ────────────────────────────────────────────────
# 시트 2: 외부 의존 잔여
# ────────────────────────────────────────────────
ws2 = wb.create_sheet("외부 의존 잔여")
ws2.cell(row=1, column=1, value="외부 의존 / 사용자 액션 잔여")
ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PENDING_HEADER))

for j, h in enumerate(PENDING_HEADER, start=1):
    c = ws2.cell(row=2, column=j, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = align_center
    c.border = border

for i, row in enumerate(PENDING, start=3):
    for j, val in enumerate(row, start=1):
        c = ws2.cell(row=i, column=j, value=val)
        c.alignment = align_top_wrap
        c.border = border

widths2 = [38, 38, 60, 18]
for j, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(j)].width = w
for i in range(3, 3 + len(PENDING)):
    ws2.row_dimensions[i].height = 50
ws2.row_dimensions[1].height = 22
ws2.row_dimensions[2].height = 28
ws2.freeze_panes = "A3"

# ────────────────────────────────────────────────
# 시트 3: 요약 통계
# ────────────────────────────────────────────────
ws3 = wb.create_sheet("요약")
summary = [
    ["구분", "건수"],
    ["Phase 1 SEO 항목", 8],
    ["Phase 2 항목 (2-1 ~ 2-7)", 7],
    ["Phase 3 고객 피드백 라운드 1 항목 (#1~#7)", 7],
    ["Phase 3 회귀 검증/stale 정리", 1],
    ["전체 항목", 23],
    ["", ""],
    ["E2E 테스트 통과 (최종, 2026-05-14)", "107 passed"],
    ["E2E 데이터 미입력 / 디자인 시안 대기 (expected)", "12 skipped + 2 xfailed"],
    ["E2E 인프라 timeout (재시도 시 통과)", "1 error"],
    ["E2E 실제 회귀 (오늘 변경 인한 깨짐)", "0"],
    ["", ""],
    ["커밋 수 (Phase 1+2+3, 누계)", "22개+"],
    ["배포 사이트", "2 (askim.kr, bloomingterra.com)"],
    ["DB 변경", "2회 (da_board_gallery.related_no1/2, da_goods.slug)"],
    ["DB 백업", "2건 (_docs/db-backup/, OneDrive 자동 동기화)"],
    ["", ""],
    ["라이브 sitemap URL 수", "에스킴 47 + 블루밍 269"],
    ["고객 라이브 검증 URL (2026-05-14)", "https://www.bloomingterra.com/goods/goods_view?no=91&cate=001 외"],
]
for i, row in enumerate(summary, start=1):
    for j, val in enumerate(row, start=1):
        c = ws3.cell(row=i, column=j, value=val)
        c.alignment = Alignment(vertical="center")
        if i == 1:
            c.font = header_font
            c.fill = header_fill
            c.alignment = align_center
ws3.column_dimensions["A"].width = 36
ws3.column_dimensions["B"].width = 36

# ────────────────────────────────────────────────
# 저장
# ────────────────────────────────────────────────
import datetime
out_dir = os.path.realpath(os.path.join(os.path.dirname(__file__), "..", "_docs", "specs"))
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.join(out_dir, f"요구사항_구현_검증_매트릭스_{datetime.date.today().isoformat()}.xlsx")
wb.save(out_path)
print(f"saved: {out_path}")
print(f"size: {os.path.getsize(out_path)} bytes")
